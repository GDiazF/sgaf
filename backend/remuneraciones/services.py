from io import BytesIO
from pathlib import Path
import unicodedata
import openpyxl
from openpyxl import Workbook


from .models import MapeoBanco, MapeoMedioPago, MapeoBancoDirecto, ValeVistaConfig


def get_bank_code_map():
    return {m.nombre.strip().upper(): m.codigo for m in MapeoBanco.objects.all()}


def get_payment_method_map():
    return {m.nombre.strip().upper(): m.codigo for m in MapeoMedioPago.objects.all()}


def get_direct_bank_code_map():
    return {m.segmento: m.codigo_completo for m in MapeoBancoDirecto.objects.all()}


def get_vale_vista_config():
    return {m.clave: m.valor for m in ValeVistaConfig.objects.all()}


def normalizar_texto(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "")).encode("ascii", "ignore").decode()
    texto = texto.replace("-", "").replace("\u00f1", "n")
    return texto


def _normalize_key(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "")).encode("ascii", "ignore").decode()
    return texto.strip().upper()


def _safe_value(row, index):
    if len(row) <= index:
        return ""
    valor = row[index].value
    return valor if valor is not None else ""


def limpiar_hoja(sheet):
    header_row = None
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        filled = [cell for cell in row if cell is not None and str(cell).strip() != ""]
        if len(filled) == 14:
            header_row = idx
            break

    if header_row and header_row > 1:
        sheet.delete_rows(1, header_row - 1)

    # Eliminar filas de totales si existen
    rows_to_delete = []
    for idx, row in list(enumerate(sheet.iter_rows(values_only=True), start=1)):
        # Check backwards might be safer if deleting while iterating, but here we collect first
        # Original code iterated reversed list:
        pass
    
    # Original logic adapted:
    for idx, row in reversed(list(enumerate(sheet.iter_rows(values_only=True), start=1))):
        if any(_normalize_key(cell).replace(" ", "") == "TOTALES" for cell in row if cell):
            sheet.delete_rows(idx)
            break


def generar_archivo_bancos(archivo):
    archivo.seek(0)
    workbook = openpyxl.load_workbook(archivo)
    sheet = workbook.active

    limpiar_hoja(sheet)
    
    bank_code_map = get_bank_code_map()
    payment_method_map = get_payment_method_map()
    direct_bank_code_map = get_direct_bank_code_map()

    nuevo_workbook = Workbook()
    sheet_nuevo = nuevo_workbook.active
    headers = ["Nombre", "Detalle", "email", "Codigo Banco", "Medio de Pago", "Glosa", "Sueldo Liquido"]
    for col, header in enumerate(headers, start=1):
        sheet_nuevo.cell(row=1, column=col).value = header
    sheet_nuevo.column_dimensions["F"].width = 25

    for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        nombre = f"{_safe_value(row, 0)}{_safe_value(row, 1)}"
        sheet_nuevo.cell(row=idx, column=1).value = nombre

        valor_e = _safe_value(row, 4)
        valor_c = _safe_value(row, 2)
        valor_d = _safe_value(row, 3)
        detalle = f"{valor_e} {valor_c} {valor_d}"
        detalle = normalizar_texto(detalle)
        sheet_nuevo.cell(row=idx, column=2).value = detalle

        sheet_nuevo.cell(row=idx, column=3).value = ""

        codigo_directo = direct_bank_code_map.get(str(_safe_value(row, 5)))
        if codigo_directo:
            sheet_nuevo.cell(row=idx, column=4).value = codigo_directo
        else:
            codigo_por_nombre = bank_code_map.get(_normalize_key(_safe_value(row, 6)))
            if codigo_por_nombre:
                sheet_nuevo.cell(row=idx, column=4).value = codigo_por_nombre

        medio_pago = payment_method_map.get(_normalize_key(_safe_value(row, 10)))
        if medio_pago:
            sheet_nuevo.cell(row=idx, column=5).value = medio_pago

        valor_glosa = str(_safe_value(row, 9) or "").replace("-", "").replace(" ", "").replace(",", ".")
        try:
            valor_glosa_numerico = int(float(valor_glosa))
        except ValueError:
            valor_glosa_numerico = None
        
        celda_glosa = sheet_nuevo.cell(row=idx, column=6)
        celda_glosa.value = valor_glosa_numerico if valor_glosa_numerico is not None else valor_glosa
        if valor_glosa_numerico is not None:
            celda_glosa.number_format = "0"

        sheet_nuevo.cell(row=idx, column=7).value = _safe_value(row, 13)

    return _save_workbook(nuevo_workbook, archivo.name)


def generar_archivo_vale_vista(archivo):
    archivo.seek(0)
    workbook = openpyxl.load_workbook(archivo)
    sheet = workbook.active

    nuevo_workbook = Workbook()
    sheet_nuevo = nuevo_workbook.active

    vv_config = get_vale_vista_config()
    val_7 = vv_config.get("VALE_VISTA_COL_7", "29")
    val_8 = vv_config.get("VALE_VISTA_COL_8", "012")
    val_9 = vv_config.get("VALE_VISTA_COL_9", "0")

    for row in sheet.iter_rows():
        fila = row[0].row if row else 1

        sheet_nuevo.cell(row=fila, column=1).value = "2"
        sheet_nuevo.cell(row=fila, column=2).value = _safe_value(row, 0)
        sheet_nuevo.cell(row=fila, column=3).value = _safe_value(row, 1)

        nombre = normalizar_texto(_safe_value(row, 4))
        sheet_nuevo.cell(row=fila, column=4).value = nombre

        app = normalizar_texto(_safe_value(row, 2))
        sheet_nuevo.cell(row=fila, column=5).value = app

        apm = normalizar_texto(_safe_value(row, 3))
        sheet_nuevo.cell(row=fila, column=6).value = apm

        sheet_nuevo.cell(row=fila, column=7).value = val_7
        sheet_nuevo.cell(row=fila, column=8).value = val_8
        sheet_nuevo.cell(row=fila, column=9).value = val_9

        monto = _safe_value(row, 13)
        sheet_nuevo.cell(row=fila, column=10).value = monto
        sheet_nuevo.cell(row=fila, column=17).value = monto
        sheet_nuevo.cell(row=fila, column=18).value = "M"

    return _save_workbook(nuevo_workbook, archivo.name)


def _save_workbook(workbook, nombre_original):
    base = Path(nombre_original).stem or "archivo"
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"procesado_{base}.xlsx"
    return buffer, filename

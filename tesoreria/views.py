import re
import io
import pdfplumber
import openpyxl
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated

def extraer_datos(texto):
    """Extrae datos específicos del texto de un PDF bancario."""
    rut_pattern = re.search(r"Rut (\d{7,8}-[\dkK])", texto)  # Beneficiario
    deposito_pattern = re.search(r"N° DEPOSITO (\d+)", texto)
    fecha_instruccion_pattern = re.search(r"instruido con fecha\s*([0-9]{2}[-/.][0-9]{2}[-/.][0-9]{4})", texto, re.IGNORECASE)
    fecha_pago_pattern = re.search(r"Fecha de Pago\s*([0-9]{2}[-/.][0-9]{2}[-/.][0-9]{4})", texto, re.IGNORECASE)
    monto_pattern = re.search(r"Monto \$\s?([\d\.]+)", texto)

    rut = rut_pattern.group(1) if rut_pattern else "No encontrado"
    deposito = deposito_pattern.group(1) if deposito_pattern else "No encontrado"

    if fecha_instruccion_pattern:
        fecha_instruccion = fecha_instruccion_pattern.group(1)
    else:
        if fecha_pago_pattern:
            fecha_instruccion = fecha_pago_pattern.group(1)
        else:
            fecha_instruccion = "No encontrada"

    if monto_pattern:
        monto_str = monto_pattern.group(1).replace(".", "")
        try:
            monto = int(monto_str)
        except ValueError:
            monto = monto_str
    else:
        monto = "No encontrado"

    return rut, deposito, fecha_instruccion, monto

class ProcesarDocumentosBancoView(APIView):
    """
    Procesa múltiples archivos PDF y genera un Excel con una pestaña por archivo.
    """
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        files = request.FILES.getlist('files')
        if not files:
            return Response({"error": "No se enviaron archivos."}, status=status.HTTP_400_BAD_REQUEST)

        # Crear Workbook de Excel en memoria
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        for f in files:
            # Lógica solicitada: Tomar solo la parte después del último "_" y quitar extensión
            base_name = f.name.rsplit('.', 1)[0]
            parts = base_name.split('_')
            title_candidate = parts[-1][:31] if parts else base_name[:31]
            
            # Limpiar caracteres prohibidos en nombres de hojas de Excel: \ / * ? : [ ]
            title_candidate = re.sub(r'[\\/*?:\[\]]', '', title_candidate)
            
            # Si el nombre queda vacío (raro) o es solo espacios
            if not title_candidate.strip():
                title_candidate = "Hoja"

            # Evitar nombres duplicados (Excel no permite dos hojas con el mismo nombre)
            final_title = title_candidate
            counter = 1
            while final_title in wb.sheetnames:
                suffix = f"({counter})"
                final_title = f"{title_candidate[:31-len(suffix)]}{suffix}"
                counter += 1
            
            ws = wb.create_sheet(title=final_title)
            ws.append(["RUT", "N° Deposito", "Fecha Instrucción", "Monto"])

            results = []
            try:
                # Abrir PDF desde el flujo de bytes en memoria
                with pdfplumber.open(f) as pdf:
                    for pagina in pdf.pages:
                        texto = pagina.extract_text()
                        if texto:
                            datos = extraer_datos(texto)
                            results.append(datos)
                
                # Escribir resultados en la hoja
                for fila in results:
                    ws.append(fila)

            except Exception as e:
                # Si un archivo falla, registrar error en la hoja pero continuar con los demás
                ws.append(["Error al procesar archivo", str(e)])

        # Guardar el Excel en un stream de bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="documentos_procesados.xlsx"'
        return response
